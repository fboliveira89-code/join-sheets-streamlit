from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os

def _last_used(ws):
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    while max_row > 1:
        if any(ws.cell(row=max_row, column=c).value is not None for c in range(1, max_col + 1)):
            break
        max_row -= 1

    while max_col > 1:
        if any(ws.cell(row=r, column=max_col).value is not None for r in range(1, max_row + 1)):
            break
        max_col -= 1

    if max_row < 1 or max_col < 1:
        return 0, 0
    return max_row, max_col

def join_sheets_like_macro_colA(
    input_path: str,
    join_sheet_name: str = "Join",
    keep_headers_each_sheet: bool = True,
    blank_row_between_blocks: bool = False,
    inplace: bool = True,
):
    if not os.path.exists(input_path):
        raise FileNotFoundError(f"Ficheiro não encontrado: {input_path}")

    wb = load_workbook(input_path, data_only=False)

    if join_sheet_name in wb.sheetnames:
        ws_old = wb[join_sheet_name]
        wb.remove(ws_old)

    ws_join = wb.create_sheet(join_sheet_name, 0)

    sheet_names = [s for s in wb.sheetnames if s != join_sheet_name]

    current_row = 1
    first_block = True

    for name in sheet_names:
        ws = wb[name]
        max_row, max_col = _last_used(ws)
        if max_row == 0 or max_col == 0:
            continue

        if not first_block and blank_row_between_blocks:
            current_row += 1

        start_src_row = 1
        if (not keep_headers_each_sheet) and (not first_block):
            start_src_row = 2

        for r in range(start_src_row, max_row + 1):
            ws_join.cell(row=current_row, column=1).value = name
            for c in range(1, max_col + 1):
                ws_join.cell(row=current_row, column=c + 1).value = ws.cell(row=r, column=c).value
            current_row += 1

        first_block = False

    max_check_rows = min(ws_join.max_row, 500)
    ws_join.column_dimensions["A"].width = 28

    for col in range(2, ws_join.max_column + 1):
        max_len = 0
        for r in range(1, max_check_rows + 1):
            v = ws_join.cell(row=r, column=col).value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws_join.column_dimensions[get_column_letter(col)].width = min(max(10, max_len + 2), 60)

    save_path = input_path if inplace else os.path.splitext(input_path)[0] + "_JOINED.xlsx"
    wb.save(save_path)
    return save_path
